# -*- coding: utf-8 -*-
# this file is released under public domain and you can use without limitations

# -------------------------------------------------------------------------
# This is a sample controller
# - index is the default action of any application
# - user is required for authentication and authorization
# - download is for downloading files uploaded in the db (does streaming)
# -------------------------------------------------------------------------

# @auth.requires_login()
def index():
    return dict()

# @auth.requires_login()
def planos_de_acao():
    # print db._lastsql
    planos = db(PLANO.concluido == 'NÃO').select(orderby=db.plano.prazo)
    # print db._lastsql
    concluidos = db(PLANO.concluido == 'SIM').select(orderby=db.plano.prazo)
    # print planos
    form = SQLFORM(PLANO, _id='form_novo_plano', submit_button='Salvar')
    if form.process().accepted:
        response.flash = "Salvo com sucesso!"
        print db._lastsql
        redirect(URL('planos_de_acao'))
        print form.vars.titulo
    elif form.errors:
        response.flash = 'Erros no preenchimento ou campo vazio!'
    # else:
    #     response.flash = 'Preencha os campos para salvar ou alterar um plano de ação!'
    # form.add_button('Cancelar', URL('index'), _class='btn btn-primary')
    button = A('Cancelar', _href='../planos_de_acao', _class="btn btn-primary", _style="border-radius: 5px;")
    return dict(planos=planos, concluidos=concluidos, form=form, button=button)


# @auth.requires_login()
def editar_plano():
    plano = db(PLANO.id == request.args(0, cast=int)).select().first()
    form = SQLFORM(PLANO, plano, submit_button='Salvar')
    if form.process().accepted:
        response.flash = 'Registro alterado com sucesso!'
        redirect(URL('planos_de_acao'))
    elif form.errors:
        response.flash = 'Erros no preenchimento ou campo vazio!'
    # else:
    #     response.flash = 'Nenhuma alteração foi realizada!'
    button = A('Cancelar', _href='../planos_de_acao', _class="btn btn-primary", _style="margin-left:20px;border-radius: 5px;")
    return dict(form=form, plano=plano, button=button)


# @auth.requires_login()
def ver_concluido():
    plano = db(PLANO.id == request.args(0, cast=int)).select().first()
    form = SQLFORM(PLANO, plano, submit_button='Salvar')
    if form.process().accepted:
        response.flash = 'Registro alterado com sucesso!'
        redirect(URL('planos_de_acao'))
    elif form.errors:
        response.flash = 'Erros no formulário!'
    # else:
    #     response.flash = 'Nenhuma alteração foi realizada!'
    button = A('Voltar', _href='../planos_de_acao', _class="btn btn-primary", _style="margin-left:20px;border-radius: 5px;")
    return dict(form=form, plano=plano, button=button)


# @auth.requires_login()
def atendimentos():
    print auth.user['username']
    atendimentos = db(ATEND).select(orderby=db.atendimentos.created_on)
    em_atendimento = db(db.status_atend.status!='CONCLUÍDO').select(join=db.atendimentos.on(
        db.status_atend.id == db.atendimentos.status), orderby=db.atendimentos.created_on)

    # em_atendimento = db(db.status_atend.status!='CONCLUÍDO').select(
    #     ATEND.created_on, ATEND.cliente, ATEND.contato, ATEND.status,join=db.atendimentos.on(
    #     db.status_atend.id == db.atendimentos.status), orderby=db.atendimentos.created_on)

    atend_conc = db(db.status_atend.status=='CONCLUÍDO').select(join=db.atendimentos.on(
        db.status_atend.id == db.atendimentos.status), orderby=db.atendimentos.created_on)
    # for row in cont_aguardando:
    #     print row.status_atend.status
    # cont_aguardando = db(ATEND.status=='3').count()
    # print cont_aguardando
    # print em_atendimento
    count = ATEND.id.count()
    # print count
    dados = db(STATEND).select(
        ATEND.status, count, join=db.atendimentos.on(
        db.status_atend.id == db.atendimentos.status), groupby=(ATEND.status))
    lista = dados.as_list()
    print lista[0]['atendimentos']['status']
    print dados
    # return dados
    # print dados

    form_novo = SQLFORM(ATEND, submit_button="Salvar")
    if form_novo.process().accepted:
        redirect(URL('atendimentos'))
    elif form_novo.errors:
        response.flash = "ERRO no form"

    return dict(atendimentos=atendimentos, atend_conc=atend_conc, em_atendimento=em_atendimento, 
                dados=dados, form_novo=form_novo)
                # cont_em_atendimento=cont_em_atendimento, cont_aguardando=cont_aguardando, 
                # cont_entrar_em_contato=cont_entrar_em_contato, cont_concluido=cont_concluido)

"""
                      <li class="li-info" role="presentation" class="cor">Em atendimento <span class="badge cor-badge">
                      {{ =cont_em_atendimento.status_atend.status }}</span></li>
                      <li class="li-info" role="presentation" class="cor">Aguardando <span class="badge cor-badge">
                      {{ =cont_aguardando.status_atend.status }}</span></li>
                      <li class="li-info" role="presentation" class="cor">Entrar em contato <span class="badge cor-badge">
                      {{ =cont_entrar_em_contato.status_atend.status }}</span></li>
"""


def novo_atend():
    form1 = SQLFORM(PLANO)
    if form.process().accepted:
        print request.vars.depto

    return dict(form1=form1)

def atendimentos_ui():
    planos = db(PLANO).select(orderby=db.plano.prazo)
    return dict(planos=planos)


def layoutit():
    return dict()

def tabela():
    return dict()


"""

>>> db.define_table('person',
                    Field('name'),
                    format='%(name)s')

>>> db.define_table('thing',
                    Field('name'),
                    Field('owner_id', 'reference person'),
                    format='%(name)s')

rows = db(db.person).select(join=db.thing.on(db.person.id == db.thing.owner_id))                    
"""



def test():
    # row = db(db.status_atend.id == db.atendimentos.status).select()
    rows = db(db.status_atend.status=='CONCLUÍDO').select(join=db.atendimentos.on(
        db.status_atend.id == db.atendimentos.status), orderby=db.atendimentos.created_on)
    for row in rows:
        # print row.cliente
        print row.status_atend.status
    return dict(rows=rows)

# @auth.requires_login()
def export_xls():
    import openpyxl, os     #'''importa openpyxl'''
    from openpyxl import load_workbook
    #'''importa funções para converter letras e números'''
    # from openpyxl.cell import get_column_letter, column_index_from_string
    

    wb = load_workbook('./applications/plano_de_acao/static/files/PlanoAmauri.xlsx')
    sheet = wb.get_sheet_by_name('Plano')       #obtém a planilha Plan1
    sheet = wb['Plano']
    # anotherSheet = wb.active
    conteudo = sheet['A9']

    print sheet['B9'].value

    print 'Nome das planilhas: ', wb.get_sheet_names()

    print 'Valor: ', conteudo.value
    print 'Tamanho máximo de linhas da planilha: ', str(sheet.max_row)
    print 'Tamanho máximo de colunas da planilha: ', str(sheet.max_column)

    # print os.listdir('./applications/plano_de_acao/static/files')

    return "Leu arquivo xls"



def teste():
    return dict()


def user():
    """
    exposes:
    http://..../[app]/default/user/login
    http://..../[app]/default/user/logout
    http://..../[app]/default/user/register
    http://..../[app]/default/user/profile
    http://..../[app]/default/user/retrieve_password
    http://..../[app]/default/user/change_password
    http://..../[app]/default/user/bulk_register
    use @auth.requires_login()
        @auth.requires_membership('group name')
        @auth.requires_permission('read','table name',record_id)
    to decorate functions that need access control
    also notice there is http://..../[app]/appadmin/manage/auth to allow administrator to manage users
    """
    return dict(form=auth())


@cache.action()
def download():
    """
    allows downloading of uploaded files
    http://..../[app]/default/download/[filename]
    """
    return response.download(request, db)


def call():
    """
    exposes services. for example:
    http://..../[app]/default/call/jsonrpc
    decorate with @services.jsonrpc the functions to expose
    supports xml, json, xmlrpc, jsonrpc, amfrpc, rss, csv
    """
    return service()


